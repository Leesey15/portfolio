import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

FONT_NAME = "Arial"
NAVY = "1F3864"
LIGHT_BLUE = "DDEBF7"
WHITE = "FFFFFF"

df = pd.read_csv("superstore_clean.csv", parse_dates=["Order Date", "Ship Date"])
n = len(df)

wb = Workbook()

# ---------------------------------------------------------------
# README
# ---------------------------------------------------------------
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 100

def title_cell(cell, text, size=16, color=NAVY):
    cell.value = text
    cell.font = Font(name=FONT_NAME, size=size, bold=True, color=color)

r = 2
title_cell(ws.cell(row=r, column=1), "Retail Sales & Profitability Analysis")
r += 1
ws.cell(row=r, column=1, value="Excel Portfolio Project").font = Font(name=FONT_NAME, size=12, italic=True, color="666666")
r += 2
lines = [
    "Author: Aseye Amenuveve Gbagbo",
    "",
    "Purpose",
    "Analyze four years (2015-2018) of retail order data to identify which regions, categories, and discount",
    "levels drive — or erode — profitability, and present the findings as a formula-driven Excel dashboard.",
    "",
    "Dataset",
    "9,994 order line items from the widely-used \"Sample Superstore\" retail dataset (US-based retailer),",
    "sourced from a public GitHub mirror for portfolio/practice use. Columns: order and ship dates, ship",
    "mode, customer segment, region/state, product category and sub-category, sales, quantity, discount,",
    "and profit.",
    "",
    "How this workbook is built",
    "- Data: the cleaned raw data as an Excel Table, with three helper columns added (Year-Month, Profit",
    "  Margin, Discount Tier) computed with formulas.",
    "- Lookup Tables: reference tables used with INDEX/MATCH and nested IF logic (illustrative regional",
    "  manager names are placeholders for demonstration only, not real staff).",
    "- Summary: pivot-style breakdowns (by Category, Region, Segment, Discount Tier, and Month) built",
    "  entirely with SUMIFS/COUNTIFS/AVERAGEIFS formulas against the Data table — recalculates if the",
    "  underlying data changes.",
    "- Dashboard: KPI cards and charts summarizing the headline findings.",
    "",
    "Key finding",
    "Profit margin falls sharply as discount increases: orders with no discount average a healthy margin,",
    "while orders discounted 30% or more are frequently sold at a loss. See the Dashboard and the",
    "Discount Tier table on the Summary sheet.",
    "",
    "Tools demonstrated: Excel Tables, SUMIFS/COUNTIFS/AVERAGEIFS, INDEX/MATCH, nested IF, LARGE-based",
    "ranking, conditional formatting, PivotChart-style dashboard design.",
]
for line in lines:
    c = ws.cell(row=r, column=1, value=line)
    if line in ("Purpose", "Dataset", "How this workbook is built", "Key finding"):
        c.font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
    else:
        c.font = Font(name=FONT_NAME, size=10.5)
    ws.row_dimensions[r].height = 15
    r += 1

# ---------------------------------------------------------------
# DATA
# ---------------------------------------------------------------
ws_d = wb.create_sheet("Data")
headers = ["Order ID", "Order Date", "Ship Date", "Ship Mode", "Segment", "Region", "State",
           "Category", "Sub-Category", "Product Name", "Sales", "Quantity", "Discount", "Profit",
           "Year-Month", "Profit Margin", "Discount Tier", "Regional Manager"]
for j, h in enumerate(headers, start=1):
    c = ws_d.cell(row=1, column=j, value=h)
    c.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for i, row in df.iterrows():
    excel_row = i + 2
    ws_d.cell(row=excel_row, column=1, value=row["Order ID"])
    ws_d.cell(row=excel_row, column=2, value=row["Order Date"].to_pydatetime()).number_format = "mm/dd/yyyy"
    ws_d.cell(row=excel_row, column=3, value=row["Ship Date"].to_pydatetime()).number_format = "mm/dd/yyyy"
    ws_d.cell(row=excel_row, column=4, value=row["Ship Mode"])
    ws_d.cell(row=excel_row, column=5, value=row["Segment"])
    ws_d.cell(row=excel_row, column=6, value=row["Region"])
    ws_d.cell(row=excel_row, column=7, value=row["State"])
    ws_d.cell(row=excel_row, column=8, value=row["Category"])
    ws_d.cell(row=excel_row, column=9, value=row["Sub-Category"])
    ws_d.cell(row=excel_row, column=10, value=row["Product Name"])
    ws_d.cell(row=excel_row, column=11, value=float(row["Sales"])).number_format = '$#,##0.00'
    ws_d.cell(row=excel_row, column=12, value=float(row["Quantity"]))
    ws_d.cell(row=excel_row, column=13, value=float(row["Discount"])).number_format = '0%'
    ws_d.cell(row=excel_row, column=14, value=float(row["Profit"])).number_format = '$#,##0.00;($#,##0.00)'
    # Year-Month helper (text, formula-based on Order Date)
    ws_d.cell(row=excel_row, column=15, value=f'=TEXT(B{excel_row},"YYYY-MM")')
    # Profit margin formula
    ws_d.cell(row=excel_row, column=16, value=f'=IFERROR(N{excel_row}/K{excel_row},0)').number_format = '0.0%'
    # Discount tier via nested IF
    ws_d.cell(row=excel_row, column=17,
              value=f'=IF(M{excel_row}=0,"No Discount",IF(M{excel_row}<0.2,"Low (1-19%)",'
                    f'IF(M{excel_row}<0.4,"Medium (20-39%)","High (40%+)")))')
    # Regional manager via INDEX/MATCH against Lookup Tables sheet
    ws_d.cell(row=excel_row, column=18,
              value=f"=INDEX('Lookup Tables'!$B$2:$B$5,MATCH(F{excel_row},'Lookup Tables'!$A$2:$A$5,0))")

last_row = n + 1
for col_letter, width in zip("ABCDEFGHIJKLMNOPQR",
                              [16, 12, 12, 14, 11, 9, 14, 15, 14, 34, 11, 10, 10, 11, 11, 12, 15, 17]):
    ws_d.column_dimensions[col_letter].width = width
ws_d.freeze_panes = "A2"

tab = Table(displayName="OrdersTable", ref=f"A1:R{last_row}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws_d.add_table(tab)

# Conditional formatting: highlight negative profit
red_fill = PatternFill("solid", fgColor="FFC7CE")
ws_d.conditional_formatting.add(
    f"N2:N{last_row}",
    CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=Font(color="9C0006"))
)

# ---------------------------------------------------------------
# LOOKUP TABLES
# ---------------------------------------------------------------
ws_l = wb.create_sheet("Lookup Tables")
ws_l.column_dimensions["A"].width = 14
ws_l.column_dimensions["B"].width = 26
ws_l.column_dimensions["D"].width = 20
ws_l.column_dimensions["E"].width = 16

ws_l["A1"] = "Region"
ws_l["B1"] = "Regional Manager (sample/placeholder)"
for c in ("A1", "B1"):
    ws_l[c].font = Font(name=FONT_NAME, bold=True, color=WHITE)
    ws_l[c].fill = PatternFill("solid", fgColor=NAVY)
regions = [("Central", "Manager A"), ("East", "Manager B"), ("South", "Manager C"), ("West", "Manager D")]
for i, (reg, mgr) in enumerate(regions, start=2):
    ws_l.cell(row=i, column=1, value=reg)
    ws_l.cell(row=i, column=2, value=mgr)
ws_l["A7"] = "Note: manager names are illustrative placeholders used to demonstrate INDEX/MATCH lookups, not real staff."
ws_l["A7"].font = Font(name=FONT_NAME, size=9, italic=True, color="808080")

ws_l["D1"] = "Discount Tier"
ws_l["E1"] = "Range"
for c in ("D1", "E1"):
    ws_l[c].font = Font(name=FONT_NAME, bold=True, color=WHITE)
    ws_l[c].fill = PatternFill("solid", fgColor=NAVY)
tiers = [("No Discount", "0%"), ("Low", "1-19%"), ("Medium", "20-39%"), ("High", "40%+")]
for i, (t, rg) in enumerate(tiers, start=2):
    ws_l.cell(row=i, column=4, value=t)
    ws_l.cell(row=i, column=5, value=rg)

wb.save("Aseye_Gbagbo_Excel_Retail_Sales_Dashboard.xlsx")
print("stage 1 saved")
