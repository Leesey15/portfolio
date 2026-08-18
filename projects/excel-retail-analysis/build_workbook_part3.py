from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

FONT_NAME = "Arial"
NAVY = "1F3864"
WHITE = "FFFFFF"
ACCENT = "2E75B6"
LIGHT = "DDEBF7"

wb = load_workbook("Aseye_Gbagbo_Excel_Retail_Sales_Dashboard.xlsx")
sm = wb["Summary"]

ws = wb.create_sheet("Dashboard", 0)  # put Dashboard first (after we reorder below)
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJKL", [3, 16, 16, 16, 16, 3, 16, 16, 16, 16, 16, 3]):
    ws.column_dimensions[col].width = w

ws.merge_cells("B2:K2")
ws["B2"] = "Retail Sales & Profitability Dashboard"
ws["B2"].font = Font(name=FONT_NAME, size=20, bold=True, color=NAVY)
ws.merge_cells("B3:K3")
ws["B3"] = "Sample Superstore dataset · 2015-2018 · 9,994 order line items"
ws["B3"].font = Font(name=FONT_NAME, size=11, italic=True, color="666666")

# ---- KPI cards (row 5-8) ----
kpis = [
    ("Total Sales", "=SUM(Summary!B3:B5)", '$#,##0', "B"),
    ("Total Profit", "=SUM(Summary!C3:C5)", '$#,##0', "D"),
    ("Overall Profit Margin", "=SUM(Summary!C3:C5)/SUM(Summary!B3:B5)", '0.0%', "F"),
    ("Order Lines", "=SUM(Summary!E3:E5)", '#,##0', "H"),
]
thin = Side(style="thin", color="B7C6E0")
for label, formula, fmt, col in kpis:
    top = f"{col}5"
    ws.merge_cells(f"{col}5:{chr(ord(col)+0)}6" if False else f"{col}5:{col}5")  # placeholder, real merge below
for label, formula, fmt, col in kpis:
    ncol = chr(ord(col) + 1)
    ws.merge_cells(f"{col}5:{ncol}5")
    ws.merge_cells(f"{col}6:{ncol}8")
    cell_label = ws[f"{col}5"]
    cell_label.value = label
    cell_label.font = Font(name=FONT_NAME, size=10.5, bold=True, color=WHITE)
    cell_label.fill = PatternFill("solid", fgColor=NAVY)
    cell_label.alignment = Alignment(horizontal="center", vertical="center")
    cell_val = ws[f"{col}6"]
    cell_val.value = formula
    cell_val.number_format = fmt
    cell_val.font = Font(name=FONT_NAME, size=22, bold=True, color=ACCENT)
    cell_val.fill = PatternFill("solid", fgColor=LIGHT)
    cell_val.alignment = Alignment(horizontal="center", vertical="center")
    for rr in (5, 6, 7, 8):
        for cc in (col, ncol):
            ws[f"{cc}{rr}"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---- Insight callout ----
ws.merge_cells("B10:K10")
ws["B10"] = ("Key finding: profit margin declines sharply as discount deepens — orders with no discount "
             "average a healthy margin, while orders discounted 40%+ are frequently sold at a loss. "
             "See the Discount Tier chart below.")
ws["B10"].font = Font(name=FONT_NAME, size=10.5, italic=True, color="9C0006")
ws.row_dimensions[10].height = 28
ws["B10"].alignment = Alignment(wrap_text=True, vertical="center")

# ---- Charts ----
# 1. Sales by Category (bar)
c1 = BarChart()
c1.title = "Sales by Category"
c1.y_axis.title = "Sales ($)"
c1.style = 10
data = Reference(sm, min_col=2, min_row=2, max_row=5)
cats = Reference(sm, min_col=1, min_row=3, max_row=5)
c1.add_data(data, titles_from_data=True)
c1.set_categories(cats)
c1.height, c1.width = 8, 14
ws.add_chart(c1, "B12")

# 2. Profit by Region (bar)
c2 = BarChart()
c2.title = "Profit by Region"
c2.y_axis.title = "Profit ($)"
c2.style = 11
data = Reference(sm, min_col=3, min_row=8, max_row=12)
cats = Reference(sm, min_col=1, min_row=9, max_row=12)
c2.add_data(data, titles_from_data=True)
c2.set_categories(cats)
c2.height, c2.width = 8, 14
ws.add_chart(c2, "G12")

# 3. Profit Margin by Discount Tier (bar) - key insight
c3 = BarChart()
c3.title = "Profit Margin (%) by Discount Tier"
c3.y_axis.title = "Profit Margin"
c3.y_axis.numFmt = '0%'
c3.style = 12
data = Reference(sm, min_col=4, min_row=21, max_row=25)
cats = Reference(sm, min_col=1, min_row=22, max_row=25)
c3.add_data(data, titles_from_data=True)
c3.set_categories(cats)
c3.height, c3.width = 8, 14
ws.add_chart(c3, "B29")

# 4. Monthly Sales Trend (line)
c4 = LineChart()
c4.title = "Monthly Sales Trend (2015-2018)"
c4.y_axis.title = "Sales ($)"
c4.style = 2
data = Reference(sm, min_col=2, min_row=61, max_row=109)
cats = Reference(sm, min_col=1, min_row=62, max_row=109)
c4.add_data(data, titles_from_data=True)
c4.set_categories(cats)
c4.height, c4.width = 8, 14
for s in c4.series:
    s.smooth = False
ws.add_chart(c4, "G29")

# 5. Top 10 Sub-Categories by Profit (bar)
c5 = BarChart()
c5.title = "Top 10 Sub-Categories by Profit"
c5.type = "bar"  # horizontal
c5.y_axis.title = "Profit ($)"
c5.style = 13
data = Reference(sm, min_col=6, min_row=48, max_row=58)
cats = Reference(sm, min_col=5, min_row=49, max_row=58)
c5.add_data(data, titles_from_data=True)
c5.set_categories(cats)
c5.height, c5.width = 8, 14
ws.add_chart(c5, "B46")

wb.save("Aseye_Gbagbo_Excel_Retail_Sales_Dashboard.xlsx")
print("dashboard added; sheets:", wb.sheetnames)
