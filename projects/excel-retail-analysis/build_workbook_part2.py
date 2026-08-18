import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule

FONT_NAME = "Arial"
NAVY = "1F3864"
WHITE = "FFFFFF"
LIGHT = "DDEBF7"

FIRST_ROW, LAST_ROW = 2, 9995  # Data rows

df = pd.read_csv("superstore_clean.csv", parse_dates=["Order Date"])
categories = sorted(df["Category"].unique())
regions = sorted(df["Region"].unique())
segments = sorted(df["Segment"].unique())
subcats = sorted(df["Sub-Category"].unique())
tiers = ["No Discount", "Low (1-19%)", "Medium (20-39%)", "High (40%+)"]
months = sorted(df["Order Date"].dt.strftime("%Y-%m").unique())

wb = load_workbook("Aseye_Gbagbo_Excel_Retail_Sales_Dashboard.xlsx")

def header_row(ws, row, headers, widths=None):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
    if widths:
        from openpyxl.utils import get_column_letter
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w

# ---------------------------------------------------------------
# SUMMARY sheet
# ---------------------------------------------------------------
ws = wb.create_sheet("Summary")
ws.sheet_view.showGridLines = False

# --- By Category (A1:E5) ---
ws["A1"] = "Sales & Profit by Category"
ws["A1"].font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
header_row(ws, 2, ["Category", "Total Sales", "Total Profit", "Profit Margin %", "Order Lines"],
           widths=[20, 15, 15, 16, 12])
for i, cat in enumerate(categories, start=3):
    ws.cell(row=i, column=1, value=cat)
    ws.cell(row=i, column=2, value=f'=SUMIFS(Data!$K${FIRST_ROW}:$K${LAST_ROW},Data!$H${FIRST_ROW}:$H${LAST_ROW},A{i})').number_format = '$#,##0'
    ws.cell(row=i, column=3, value=f'=SUMIFS(Data!$N${FIRST_ROW}:$N${LAST_ROW},Data!$H${FIRST_ROW}:$H${LAST_ROW},A{i})').number_format = '$#,##0'
    ws.cell(row=i, column=4, value=f'=IFERROR(C{i}/B{i},0)').number_format = '0.0%'
    ws.cell(row=i, column=5, value=f'=COUNTIFS(Data!$H${FIRST_ROW}:$H${LAST_ROW},A{i})')
cat_end = 2 + len(categories)

# --- By Region (A8:E12) ---
r0 = cat_end + 2
ws.cell(row=r0, column=1, value="Sales & Profit by Region").font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
header_row(ws, r0 + 1, ["Region", "Total Sales", "Total Profit", "Profit Margin %", "Order Lines"])
for i, reg in enumerate(regions, start=r0 + 2):
    ws.cell(row=i, column=1, value=reg)
    ws.cell(row=i, column=2, value=f'=SUMIFS(Data!$K${FIRST_ROW}:$K${LAST_ROW},Data!$F${FIRST_ROW}:$F${LAST_ROW},A{i})').number_format = '$#,##0'
    ws.cell(row=i, column=3, value=f'=SUMIFS(Data!$N${FIRST_ROW}:$N${LAST_ROW},Data!$F${FIRST_ROW}:$F${LAST_ROW},A{i})').number_format = '$#,##0'
    ws.cell(row=i, column=4, value=f'=IFERROR(C{i}/B{i},0)').number_format = '0.0%'
    ws.cell(row=i, column=5, value=f'=COUNTIFS(Data!$F${FIRST_ROW}:$F${LAST_ROW},A{i})')
reg_end = r0 + 1 + len(regions)

# --- By Segment ---
r1 = reg_end + 2
ws.cell(row=r1, column=1, value="Sales & Profit by Segment").font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
header_row(ws, r1 + 1, ["Segment", "Total Sales", "Total Profit", "Profit Margin %", "Order Lines"])
for i, seg in enumerate(segments, start=r1 + 2):
    ws.cell(row=i, column=1, value=seg)
    ws.cell(row=i, column=2, value=f'=SUMIFS(Data!$K${FIRST_ROW}:$K${LAST_ROW},Data!$E${FIRST_ROW}:$E${LAST_ROW},A{i})').number_format = '$#,##0'
    ws.cell(row=i, column=3, value=f'=SUMIFS(Data!$N${FIRST_ROW}:$N${LAST_ROW},Data!$E${FIRST_ROW}:$E${LAST_ROW},A{i})').number_format = '$#,##0'
    ws.cell(row=i, column=4, value=f'=IFERROR(C{i}/B{i},0)').number_format = '0.0%'
    ws.cell(row=i, column=5, value=f'=COUNTIFS(Data!$E${FIRST_ROW}:$E${LAST_ROW},A{i})')
seg_end = r1 + 1 + len(segments)

# --- By Discount Tier (the headline insight) ---
r2 = seg_end + 2
ws.cell(row=r2, column=1, value="Profit Margin by Discount Tier (key finding)").font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
header_row(ws, r2 + 1, ["Discount Tier", "Total Sales", "Total Profit", "Profit Margin %", "Order Lines"])
for i, t in enumerate(tiers, start=r2 + 2):
    ws.cell(row=i, column=1, value=t)
    ws.cell(row=i, column=2, value=f'=SUMIFS(Data!$K${FIRST_ROW}:$K${LAST_ROW},Data!$Q${FIRST_ROW}:$Q${LAST_ROW},A{i})').number_format = '$#,##0'
    ws.cell(row=i, column=3, value=f'=SUMIFS(Data!$N${FIRST_ROW}:$N${LAST_ROW},Data!$Q${FIRST_ROW}:$Q${LAST_ROW},A{i})').number_format = '$#,##0'
    ws.cell(row=i, column=4, value=f'=IFERROR(C{i}/B{i},0)').number_format = '0.0%'
    ws.cell(row=i, column=5, value=f'=COUNTIFS(Data!$Q${FIRST_ROW}:$Q${LAST_ROW},A{i})')
tier_end = r2 + 1 + len(tiers)

# --- Top 10 Sub-Categories by Profit ---
r3 = tier_end + 2
ws.cell(row=r3, column=1, value="Sub-Category Profit (all 17, for ranking)").font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
header_row(ws, r3 + 1, ["Sub-Category", "Total Profit"], widths=None)
for i, sc in enumerate(subcats, start=r3 + 2):
    ws.cell(row=i, column=1, value=sc)
    ws.cell(row=i, column=2, value=f'=SUMIFS(Data!$N${FIRST_ROW}:$N${LAST_ROW},Data!$I${FIRST_ROW}:$I${LAST_ROW},A{i})').number_format = '$#,##0'
subcat_start, subcat_end = r3 + 2, r3 + 1 + len(subcats)

r4 = subcat_end + 2
ws.cell(row=r4, column=4, value="Top 10 Sub-Categories by Profit").font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
header_row(ws, r4 + 1, ["", "", "", "Rank", "Sub-Category", "Total Profit"])
for k in range(1, 11):
    row = r4 + 1 + k
    ws.cell(row=row, column=4, value=k)
    ws.cell(row=row, column=6, value=f'=LARGE($B${subcat_start}:$B${subcat_end},D{row})').number_format = '$#,##0'
    ws.cell(row=row, column=5,
            value=f'=INDEX($A${subcat_start}:$A${subcat_end},MATCH(F{row},$B${subcat_start}:$B${subcat_end},0))')
top10_end = r4 + 1 + 10

# --- Monthly trend ---
r5 = top10_end + 2
ws.cell(row=r5, column=1, value="Monthly Sales & Profit Trend (2015-2018)").font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
header_row(ws, r5 + 1, ["Year-Month", "Total Sales", "Total Profit"])
for i, ym in enumerate(months, start=r5 + 2):
    ws.cell(row=i, column=1, value=ym)
    ws.cell(row=i, column=2, value=f'=SUMIFS(Data!$K${FIRST_ROW}:$K${LAST_ROW},Data!$O${FIRST_ROW}:$O${LAST_ROW},A{i})').number_format = '$#,##0'
    ws.cell(row=i, column=3, value=f'=SUMIFS(Data!$N${FIRST_ROW}:$N${LAST_ROW},Data!$O${FIRST_ROW}:$O${LAST_ROW},A{i})').number_format = '$#,##0'
month_start, month_end = r5 + 2, r5 + 1 + len(months)

ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 15

# conditional color scale on discount-tier margin column to visualize the finding
ws.conditional_formatting.add(
    f"D{r2+2}:D{tier_end}",
    ColorScaleRule(start_type="min", start_color="F8696B", end_type="max", end_color="63BE7B")
)

wb.save("Aseye_Gbagbo_Excel_Retail_Sales_Dashboard.xlsx")
print("summary ranges:", dict(
    cat=(3, cat_end), reg=(r0+2, reg_end), seg=(r1+2, seg_end), tier=(r2+2, tier_end),
    top10=(r4+2, top10_end), month=(month_start, month_end)
))
